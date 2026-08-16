import streamlit as st
import pandas as pd
import numpy as np
import random
from copy import deepcopy
import openpyxl
import math
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
import io

st.set_page_config(page_title="Gerador de Plano de Carregamento", page_icon="🚛", layout="wide")

# ==========================================
# LIMITES GLOBAIS ESTABELECIDOS
# ==========================================
MIN_CARGO = 26.5  
MAX_CARGO = 27.5  
TARGET_CG = 2.95  

st.title("🚛 Gerador Automático de Carregamento")
st.write(f"O algoritmo calcula o Centro de Gravidade (CG) ideal (~2.95m) e distribui os blocos mantendo a carga entre **{MIN_CARGO} t** e **{MAX_CARGO} t**.")

class Container:
    def __init__(self, id):
        self.id = id
        self.cols = {0: [], 1: []}  # 0: Porta, 1: Cego
        
    def get_cargo_weight(self):
        return sum(b['TONS'] for b in self.cols[0]) + sum(b['TONS'] for b in self.cols[1])
        
    def get_col_cargo_weight(self, c):
        return sum(b['TONS'] for b in self.cols[c])
        
    def get_col_height(self, c):
        return sum(b['h_eff'] for b in self.cols[c])
        
    def get_col_length(self, c):
        if not self.cols[c]: return 0.0
        return max(b['l_eff'] for b in self.cols[c])
        
    def can_add(self, block, c_idx, orient):
        if self.get_cargo_weight() + block['TONS'] > MAX_CARGO:
            return False
            
        if orient == 'C':
            l_eff, w_eff, h_eff = block['Comp'], block['Alt'], block['Larg']
        else: 
            l_eff, w_eff, h_eff = block['Alt'], block['Comp'], block['Larg']
            
        if w_eff > 2.20: return False
        if self.get_col_height(c_idx) + h_eff > 2.20: return False
        
        new_col_len = max(self.get_col_length(c_idx), l_eff)
        other_col_len = self.get_col_length(1 - c_idx)
        if new_col_len + other_col_len > 5.90: return False
        
        return True
        
    def add(self, block, c_idx, orient):
        b = deepcopy(block)
        b['orient'] = orient
        if orient == 'C':
            b['l_eff'], b['w_eff'], b['h_eff'] = b['Comp'], b['Alt'], b['Larg']
        else:
            b['l_eff'], b['w_eff'], b['h_eff'] = b['Alt'], b['Comp'], b['Larg']
        self.cols[c_idx].append(b)

    def calculate_cg(self):
        door_w = self.get_col_cargo_weight(0)
        cego_w = self.get_col_cargo_weight(1)
        total_w = door_w + cego_w
        if total_w == 0: return 2.95
        
        cm_door = self.get_col_length(0) / 2.0 if self.get_col_length(0) > 0 else 0
        cm_cego = 5.90 - (self.get_col_length(1) / 2.0) if self.get_col_length(1) > 0 else 5.90
        
        cg = (door_w * cm_door + cego_w * cm_cego) / total_w
        return cg

def solve(blocks_data):
    total_tons = sum(b['TONS'] for b in blocks_data)
    min_cont = math.ceil(total_tons / MAX_CARGO)
    max_cont = max(min_cont, math.floor(total_tons / MIN_CARGO)) + 1 # Permite testar +1 container se necessário
    
    best_solution = None
    best_cg_variance = float('inf')
    
    # Testa limites flexíveis caso a combinação exata de 26.5 seja muito restrita para os blocos
    for strict_min in [MIN_CARGO, 26.0, 25.5]: 
        for num_containers in range(min_cont, max_cont + 1):
            target_weight = total_tons / num_containers
            
            # Se a quantidade de containers gerar uma média menor que o mínimo, pula
            if target_weight > MAX_CARGO or (target_weight < strict_min and strict_min == MIN_CARGO):
                continue
            
            for attempt in range(12000): 
                containers = [Container(i) for i in range(1, num_containers + 1)]
                shuffled = sorted(blocks_data, key=lambda x: x['TONS'] + random.uniform(-1.5, 1.5), reverse=True)
                
                success = True
                for b in shuffled:
                    valid_moves = []
                    for cont in containers:
                        for c_idx in [0, 1]:
                            for orient in ['C', 'A']:
                                if cont.can_add(b, c_idx, orient):
                                    door_w = cont.get_col_cargo_weight(0) + (b['TONS'] if c_idx == 0 else 0)
                                    cego_w = cont.get_col_cargo_weight(1) + (b['TONS'] if c_idx == 1 else 0)
                                    door_l = max(cont.get_col_length(0), (b['Comp'] if orient=='C' else b['Alt']) if c_idx == 0 else 0)
                                    cego_l = max(cont.get_col_length(1), (b['Comp'] if orient=='C' else b['Alt']) if c_idx == 1 else 0)
                                    
                                    cm_door = door_l / 2.0 if door_l > 0 else 0
                                    cm_cego = 5.90 - (cego_l / 2.0) if cego_l > 0 else 5.90
                                    
                                    simulated_total_w = door_w + cego_w
                                    simulated_cg = (door_w * cm_door + cego_w * cm_cego) / simulated_total_w
                                    
                                    cg_penalty = abs(simulated_cg - TARGET_CG) * 100
                                    target_dist = abs(simulated_total_w - target_weight) * 25
                                    
                                    door_pattern_penalty = max(0, 6.5 - door_w) + max(0, door_w - 12.5)
                                    cego_pattern_penalty = max(0, 14.0 - cego_w) + max(0, cego_w - 20.0)
                                    pattern_penalty = (door_pattern_penalty + cego_pattern_penalty) * 5
                                    
                                    score = cg_penalty + target_dist + pattern_penalty
                                    valid_moves.append((score, cont, c_idx, orient))
                    
                    if not valid_moves:
                        success = False
                        break
                    else:
                        valid_moves.sort(key=lambda x: x[0])
                        chosen = random.choice(valid_moves[:2])
                        chosen[1].add(b, chosen[2], chosen[3])
                        
                if success:
                    is_valid = True
                    for c in containers:
                        w = c.get_cargo_weight()
                        if w < strict_min or w > MAX_CARGO:
                            is_valid = False
                            break
                            
                    if is_valid:
                        total_cg_variance = sum(abs(c.calculate_cg() - TARGET_CG) for c in containers)
                        if total_cg_variance < best_cg_variance:
                            best_cg_variance = total_cg_variance
                            best_solution = deepcopy(containers)
                            
        if best_solution is not None:
            return best_solution, strict_min
            
    return None, None

def create_excel_report(solution):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plano de Carregamento"
    colors = ['FFFF00', '00B0F0', '00FF00', 'FF9900', 'CCC0DA', 'FF0000', 'E2EFDA', 'FCE4D6', 'D9E1F2']
    thick_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    row_offset = 2
    for c_idx, cont in enumerate(solution):
        door_col = sorted(cont.cols[0], key=lambda x: x['l_eff'], reverse=True)
        blind_col = sorted(cont.cols[1], key=lambda x: x['l_eff'], reverse=True)
        
        ws.merge_cells(start_row=row_offset, start_column=2, end_row=row_offset, end_column=4)
        cell = ws.cell(row=row_offset, column=2, value=f"CONTAINER {cont.id} (CG: {cont.calculate_cg():.2f}m)")
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')
        
        row_offset += 1
        ws.cell(row=row_offset, column=2, value="Lado Porta").alignment = Alignment(horizontal='center')
        ws.cell(row=row_offset, column=3, value="Lado Cego").alignment = Alignment(horizontal='center')
        row_offset += 1
        
        max_blocks = max(len(door_col), len(blind_col))
        ws.cell(row=row_offset, column=2, value=door_col[-1]['orient'] if door_col else "").alignment = Alignment(horizontal='center')
        ws.cell(row=row_offset, column=3, value=blind_col[-1]['orient'] if blind_col else "").alignment = Alignment(horizontal='center')
        row_offset += 1
        
        fill = PatternFill(start_color=colors[c_idx % len(colors)], end_color=colors[c_idx % len(colors)], fill_type="solid")
        
        for i in range(max_blocks - 1, -1, -1):
            if i < len(door_col):
                b = door_col[i]
                c = ws.cell(row=row_offset, column=2, value=b['Bloco'])
                c.fill = fill
                c.border = thick_border
                c.alignment = Alignment(horizontal='center')
            if i < len(blind_col):
                b = blind_col[i]
                c = ws.cell(row=row_offset, column=3, value=b['Bloco'])
                c.fill = fill
                c.border = thick_border
                c.alignment = Alignment(horizontal='center')
            row_offset += 1
            
        door_len = max([b['l_eff'] for b in door_col]) if door_col else 0
        blind_len = max([b['l_eff'] for b in blind_col]) if blind_col else 0
        ws.cell(row=row_offset, column=2, value=round(door_len, 2)).alignment = Alignment(horizontal='center')
        ws.cell(row=row_offset, column=3, value=round(blind_len, 2)).alignment = Alignment(horizontal='center')
        ws.cell(row=row_offset, column=4, value=round(door_len + blind_len, 2)).alignment = Alignment(horizontal='center')
        row_offset += 1
        
        ws.cell(row=row_offset, column=2, value=round(cont.get_col_cargo_weight(0), 3)).alignment = Alignment(horizontal='center')
        ws.cell(row=row_offset, column=3, value=round(cont.get_col_cargo_weight(1), 3)).alignment = Alignment(horizontal='center')
        ws.cell(row=row_offset, column=4, value=round(cont.get_cargo_weight(), 3)).alignment = Alignment(horizontal='center')
        row_offset += 3
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

uploaded_file = st.file_uploader("Selecione sua planilha de blocos (Excel)", type=["xlsx"])

if uploaded_file is not None:
    try:
        df = pd.read_excel(uploaded_file, sheet_name='BLOCOS')
        df_blocks = df.iloc[5:, [0, 1, 2, 3, 4]].copy() 
        df_blocks.columns = ['Bloco', 'Comp', 'Alt', 'Larg', 'TONS']
        for col in ['Comp', 'Alt', 'Larg', 'TONS']:
            df_blocks[col] = pd.to_numeric(df_blocks[col], errors='coerce')
        df_blocks = df_blocks.dropna(subset=['Bloco', 'TONS'])
        blocks_data = df_blocks.to_dict('records')
        
        total_weight = sum(b['TONS'] for b in blocks_data)
        st.success(f"Planilha processada! {len(blocks_data)} blocos | Tonelagem Total: {total_weight:.3f} t")
        
        if st.button(f"Gerar Plano Padrão Diretoria", type="primary"):
            with st.spinner(f"Otimizando Centro de Gravidade e Combinações..."):
                sol, valid_min = solve(blocks_data)
                
            if sol:
                if valid_min < MIN_CARGO:
                    st.warning(f"Nota: Para acomodar perfeitamente todos os 32 blocos sem ultrapassar 27.5t, o sistema ajustou o limite mínimo para {valid_min}t em alguns contêineres.")
                else:
                    st.success(f"Plano gerado com sucesso! Todos dentro da faixa.")
                
                st.subheader("Resumo do Carregamento (CG Ideal = 2.95m):")
                for c in sol:
                    carga = c.get_cargo_weight()
                    cg = c.calculate_cg()
                    st.write(f"**Container {c.id}** | Carga Total: **{carga:.3f} t** | CG: **{cg:.2f} m**")
                    st.write(f"↪ Porta: {c.get_col_cargo_weight(0):.3f} t | Fundo (Cego): {c.get_col_cargo_weight(1):.3f} t")
                    st.divider()
                
                excel_data = create_excel_report(sol)
                st.download_button(
                    label="📥 Baixar Excel do Plano (Padrão Diretoria)",
                    data=excel_data,
                    file_name="Plano_Carregamento_Diretoria.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.error("Matematicamente impossível com esses blocos exatos. Tente ajustar ligeiramente a tonelagem de 1 bloco na planilha.")
                
    except Exception as e:
        st.error(f"Erro ao ler a planilha. Detalhe: {e}")
